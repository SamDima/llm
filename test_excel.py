"""Test model's ability to work with Excel data."""

import os
import sys
import openpyxl
from openai import OpenAI

BASE_URL = sys.argv[1] if len(sys.argv) > 1 else "http://localhost:8000/v1"
EXCEL_FILE = sys.argv[2] if len(sys.argv) > 2 else None
API_KEY = os.environ.get("RUNPOD_API_KEY", "not-needed")

if not EXCEL_FILE:
    print("Usage: python test_excel.py <RUNPOD_URL> <excel_file.xlsx>")
    print("Example: python test_excel.py http://your-pod-ip:8000/v1 data.xlsx")
    sys.exit(1)

# Read Excel file
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

sheets_data = {}
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(cell) if cell is not None else "" for cell in row])
    sheets_data[sheet_name] = rows

MAX_ROWS = int(sys.argv[3]) if len(sys.argv) > 3 else 10
MAX_CHARS = 5000  # ~1500 tokens

client = OpenAI(base_url=BASE_URL, api_key=API_KEY)
models = client.models.list()

print(f"Loaded {len(sheets_data)} sheets: {list(sheets_data.keys())}")
print(f"Sending first {MAX_ROWS} rows per sheet, one sheet at a time.\n")

for sheet_name, rows in sheets_data.items():
    excel_text = f"=== Sheet: {sheet_name} ({len(rows)} rows, {len(rows[0]) if rows else 0} cols) ===\n"
    for row in rows[:MAX_ROWS]:
        line = " | ".join(row)
        excel_text += line[:300] + "\n"  # truncate wide rows
    if len(rows) > MAX_ROWS:
        excel_text += f"... ({len(rows) - MAX_ROWS} more rows)\n"

    # Trim to max chars
    excel_text = excel_text[:MAX_CHARS]

    print(f"\n--- Analyzing sheet: {sheet_name} ({len(excel_text)} chars) ---")

    response = client.chat.completions.create(
        model=models.data[0].id,
        messages=[
            {
                "role": "system",
                "content": "You are a data analyst. Analyze the provided Excel data carefully. Respond in Russian.",
            },
            {
                "role": "user",
                "content": f"Here is the Excel data:\n{excel_text}\n\nКратко опиши: что за данные, структура, ключевые поля.",
            },
        ],
        max_tokens=1024,
        temperature=0.3,
    )

    print(response.choices[0].message.content)
